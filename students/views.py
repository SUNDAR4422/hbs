from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from openpyxl import load_workbook
from .models import Student, Department, AcademicYear
from .serializers import StudentSerializer, DepartmentSerializer, BulkStudentUploadSerializer
from audit.utils import log_activity
import io


class StudentProfileView(generics.RetrieveUpdateAPIView):
    """Get and update student profile."""
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_object(self):
        return self.request.user.student_profile


class DepartmentListView(generics.ListAPIView):
    """List all departments."""
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]


class BulkStudentUploadView(APIView):
    """Upload multiple students via Excel file (Dean only)."""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Only dean can upload students
        if not request.user.is_dean() and not request.user.is_superuser:
            return Response(
                {'error': 'Only Dean can upload students'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file uploaded'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validate file extension
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'File must be an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Read Excel file
            workbook = load_workbook(io.BytesIO(file.read()))
            sheet = workbook.active
            
            students_data = []
            errors = []
            
            # Skip header row
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                try:
                    # Handle date_of_birth - convert datetime to string if needed
                    date_value = row[2]
                    if hasattr(date_value, 'strftime'):  # datetime or date object
                        date_str = date_value.strftime('%d/%m/%Y')
                    else:
                        date_str = str(date_value).strip()
                    
                    # Safely convert year fields to int
                    try:
                        current_year = int(row[6])
                    except (ValueError, TypeError):
                        raise ValueError(f"Current year must be a number, got '{row[6]}'")
                    
                    try:
                        admission_year = int(row[7])
                    except (ValueError, TypeError):
                        raise ValueError(f"Admission year must be a number, got '{row[7]}'")
                    
                    try:
                        graduation_year = int(row[8])
                    except (ValueError, TypeError):
                        raise ValueError(f"Graduation year must be a number, got '{row[8]}'")
                    
                    student_data = {
                        'register_number': str(row[0]).strip(),
                        'name': str(row[1]).strip(),
                        'date_of_birth': date_str,
                        'gender': str(row[3]).strip().upper(),
                        'department_code': str(row[4]).strip().upper(),
                        'degree': str(row[5]).strip(),
                        'current_year': current_year,
                        'admission_year': admission_year,
                        'graduation_year': graduation_year,
                        'hostel_code': str(row[9]).strip().upper() if row[9] else '',
                        'phone_number': str(row[10]).strip() if row[10] else '',
                        'email': str(row[11]).strip(),
                    }
                    students_data.append((idx, student_data))
                except Exception as e:
                    errors.append(f"Row {idx}: Error parsing data - {str(e)}")

            if errors:
                return Response(
                    {'error': 'File parsing errors', 'details': errors},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate and create students
            created_students = []
            validation_errors = []
            
            with transaction.atomic():
                for row_num, student_data in students_data:
                    serializer = BulkStudentUploadSerializer(data=student_data)
                    
                    if serializer.is_valid():
                        try:
                            student = serializer.create_student()
                            created_students.append(student.register_number)
                        except Exception as e:
                            validation_errors.append(f"Row {row_num}: {str(e)}")
                    else:
                        validation_errors.append(f"Row {row_num}: {serializer.errors}")
                
                if validation_errors:
                    transaction.set_rollback(True)
                    return Response(
                        {'error': 'Validation errors', 'details': validation_errors},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            log_activity(
                request.user,
                'BULK_STUDENT_UPLOAD',
                f'Uploaded {len(created_students)} students'
            )
            
            return Response({
                'message': f'Successfully created {len(created_students)} student accounts',
                'created': created_students
            }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class StudentListView(generics.ListAPIView):
    """List all students (Dean and Warden only)."""
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        if user.is_dean() or user.is_superuser:
            return Student.objects.all()
        elif user.is_warden():
            # Warden can only see students in their hostel
            try:
                warden_profile = user.warden_profile
                return Student.objects.filter(hostel=warden_profile.hostel)
            except:
                return Student.objects.none()
        else:
            return Student.objects.none()


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Get, update, or delete a student (Dean only)."""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, *args, **kwargs):
        if not request.user.is_dean() and not request.user.is_superuser:
            return Response(
                {'error': 'Only Dean can update students'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        response = super().update(request, *args, **kwargs)
        log_activity(request.user, 'UPDATE_STUDENT', f'Updated student: {self.get_object().register_number}')
        return response

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_dean() and not request.user.is_superuser:
            return Response(
                {'error': 'Only Dean can delete students'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        student = self.get_object()
        register_number = student.register_number
        
        # Delete associated user account
        user = student.user
        response = super().destroy(request, *args, **kwargs)
        user.delete()
        
        log_activity(request.user, 'DELETE_STUDENT', f'Deleted student: {register_number}')
        return response


class StudentCreateView(generics.CreateAPIView):
    """Create a new student (Dean only)."""
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if not request.user.is_dean() and not request.user.is_superuser:
            return Response(
                {'error': 'Only Dean can create students'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Create user account first
        from accounts.models import User
        from django.contrib.auth.hashers import make_password
        from datetime import datetime
        
        try:
            # Parse date_of_birth
            dob_str = request.data.get('date_of_birth')
            if '/' in dob_str:
                dob = datetime.strptime(dob_str, '%d/%m/%Y').date()
            else:
                dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
            
            # Create user
            name_parts = request.data.get('name', '').split()
            first_name = name_parts[0] if name_parts else ''
            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
            
            user = User.objects.create(
                username=request.data.get('register_number'),
                password=make_password(dob.strftime('%d/%m/%Y')),
                first_name=first_name,
                last_name=last_name,
                email=request.data.get('email'),
                role='student',
                must_change_password=True
            )
            
            # Create student profile
            student = Student.objects.create(
                user=user,
                register_number=request.data.get('register_number'),
                name=request.data.get('name'),
                date_of_birth=dob,
                gender=request.data.get('gender'),
                department_id=request.data.get('department'),
                degree=request.data.get('degree'),
                current_year=request.data.get('current_year'),
                admission_year=request.data.get('admission_year'),
                graduation_year=request.data.get('graduation_year'),
                hostel_id=request.data.get('hostel') if request.data.get('hostel') else None,
                phone_number=request.data.get('phone_number', ''),
            )
            
            log_activity(request.user, 'CREATE_STUDENT', f'Created student: {student.register_number}')
            
            serializer = self.get_serializer(student)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

class DepartmentManageView(generics.ListCreateAPIView):
    """List and create departments (Dean only)."""
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Department.objects.all()
    
    def perform_create(self, serializer):
        if not self.request.user.is_dean() and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only Dean can create departments')
        
        department = serializer.save()
        log_activity(
            self.request.user,
            'CREATE',
            f'Created department: {department.code} - {department.name}',
            self.request
        )


class DepartmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update or delete department (Dean only)."""
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    queryset = Department.objects.all()
    
    def perform_update(self, serializer):
        if not self.request.user.is_dean() and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only Dean can update departments')
        
        department = serializer.save()
        log_activity(
            self.request.user,
            'UPDATE',
            f'Updated department: {department.code}',
            self.request
        )
    
    def perform_destroy(self, instance):
        if not self.request.user.is_dean() and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only Dean can delete departments')
        
        log_activity(
            self.request.user,
            'DELETE',
            f'Deleted department: {instance.code}',
            self.request
        )
        instance.delete()


class AcademicYearView(APIView):
    """Get and update current academic year (Dean only)."""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get current academic year."""
        academic_year = AcademicYear.get_current()
        return Response({
            'id': academic_year.id,
            'current_year': academic_year.current_year,
            'display': f"{academic_year.current_year}-{academic_year.current_year + 1}",
            'updated_at': academic_year.updated_at
        })
    
    def put(self, request):
        """Update academic year and student years (Dean only)."""
        if not (request.user.is_superuser or hasattr(request.user, 'dean_profile')):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only Dean can update academic year')
        
        new_year = request.data.get('current_year')
        if not new_year:
            return Response(
                {'error': 'current_year is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            new_year = int(new_year)
        except ValueError:
            return Response(
                {'error': 'current_year must be a valid integer'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        academic_year = AcademicYear.get_current()
        old_year = academic_year.current_year
        
        with transaction.atomic():
            # Update academic year
            academic_year.current_year = new_year
            academic_year.updated_by = request.user
            academic_year.save()
            
            # Update all students' current_year
            students = Student.objects.all()
            updated_count = 0
            
            for student in students:
                # Calculate new current year based on admission year
                years_since_admission = new_year - student.admission_year + 1
                new_current_year = min(max(years_since_admission, 1), 4)
                
                if student.current_year != new_current_year:
                    student.current_year = new_current_year
                    student.save(update_fields=['current_year', 'updated_at'])
                    updated_count += 1
            
            log_activity(
                request.user,
                'UPDATE',
                f'Updated academic year from {old_year} to {new_year}. Updated {updated_count} students.',
                request
            )
            
            return Response({
                'message': f'Academic year updated to {new_year}-{new_year + 1}',
                'students_updated': updated_count,
                'current_year': new_year,
                'display': f"{new_year}-{new_year + 1}"
            })

